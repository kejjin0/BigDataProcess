#!/usr/bin/python3

from openpyxl import load_workbook
import math

wb=load_workbook(filename='student.xlsx')
ws=wb.active

rangeE=wb['Sheet1']

########### total 구하고 저장
count=0
all={}
totalList= []


aList=[]
bList=[]
cList=[]
fList=[]

for i in range(2,76):
	total =0
	for j in range(67,71):
		find = chr(j) + str(i)
		if j==67:
			multi=0.3
		elif j==68:
			multi=0.35
		elif j==69:
			multi=0.34
		else:
			multi=1
		total = total + (rangeE[find].value *multi )
		if multi==1 and rangeE[find].value == 0:
			total = 0

				
	find = chr(71) + str(i)
	ws[find] = total
	count += 1
	stu = 'A' + str(i)
	allKey = rangeE[stu].value
	all[allKey]=total
	totalList.append(total)

totalList.sort(reverse=True)




gradeA = math.trunc( count * 0.3 )
gradeB = math.trunc ( count * 0.7 )





i=0
for total in totalList:
        if total <= 40:
                fList.append(total)
        elif i < gradeA:
                aList.append(total)
                i = i + 1
        elif i < gradeB:
                bList.append(total)
                i = i + 1
        else:
                cList.append(total)




for total in reversed(aList):
	if bList[0] == total:
		bList.append(total)
		aList.remove(total)
	elif bList[0] < total:
		break
bList.sort(reverse=True)

for total in reversed(bList):
	if cList[0] == total:
		cList.append(total)
		bList.remove(total)
	elif cList[0] < total:
		break
cList.sort(reverse=True)




plusACount = len(aList) // 2
plusA=[]
zeroA=[]

i=0
for total in aList:
	if i < plusACount:
		plusA.append(total)
	else:
		zeroA.append(total)
	i = i + 1

for total in reversed(plusA):
	if zeroA[0] == total:
		zeroA.append(total)
		plusA.remove(total)
	elif zeroA[0] < total:
		break




plusBCount = len(bList) // 2
plusB=[]
zeroB=[]

i=0
for total in bList:
	if i < plusBCount:
		plusB.append(total)
	else:
		zeroB.append(total)
	i = i + 1

for total in reversed(plusB):
	if zeroB[0] == total:
		zeroB.append(total)
		plusB.remove(total)
	elif zeroB[0] < total:
		break



plusCCount = len(cList) // 2
plusC=[]
zeroC=[]

i=0
for total in cList:
	if i < plusCCount:
		plusC.append(total)
	else:
		zeroC.append(total)
	i = i + 1

for total in reversed(plusC):
	if zeroC[0] == total:
		zeroC.append(total)
		plusC.remove(total)
	elif zeroC[0] < total:
		break





for i in range(2,76):
	totalRange = chr(71) + str(i)
	gradeRange = chr(72) + str(i)
	stuTotal = rangeE[totalRange].value

	if stuTotal in aList:
		if stuTotal in plusA:
			ws[gradeRange] = 'A+'
		else:
			ws[gradeRange] = 'A0'

	elif stuTotal in bList:
		if stuTotal in plusB:
			ws[gradeRange] = 'B+'
		else:
			ws[gradeRange] = 'B0'

	elif stuTotal in cList:
		if stuTotal  in plusC:
			ws[gradeRange] = 'C+'
		else:
			ws[gradeRange] = 'C0'

	else:
		ws[gradeRange] = 'F'


wb.save("student.xlsx")
